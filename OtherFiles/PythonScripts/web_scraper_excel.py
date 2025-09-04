import os
import sys
import time
import subprocess
import ctypes
import importlib.util
import pandas as pd
import pyautogui
import pyperclip
import pygetwindow as gw
from tkinter import messagebox, Tk, simpledialog
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from tqdm import tqdm
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font as XLFont
import re
import tkinter as tk
from tkinter import ttk, messagebox
from tkinter.font import Font

class HaryanaEBillingScraper:
    def __init__(self, website_url):
        self.website_url = "https://works.haryana.gov.in/HEWP_Login/login.aspx"
        self.driver = None
        self.wait = None
        self.all_data = []
        self.wait_time = 15
        self.max_retries = 3
        self._chrome_was_launched = False
        self.notification_root = None


        self._hide_console()

        if self._check_prerequisites():
            print("[INIT] Prerequisites OK. Connecting to browser...")
            self.connect_to_browser()
            self._hide_console()

    def _show_console(self):
        if sys.platform == 'win32':
            print("[CONSOLE] Showing console window.")
            ctypes.windll.user32.ShowWindow(ctypes.windll.kernel32.GetConsoleWindow(), 1)

    def _hide_console(self):
        if sys.platform == 'win32':
            print("[CONSOLE] Hiding console window.")
            ctypes.windll.user32.ShowWindow(ctypes.windll.kernel32.GetConsoleWindow(), 0)

    def _check_prerequisites(self):
        """Verify all requirements are met"""
        print("[CHECK] Checking prerequisites...")
        print("="*50)
        print("CHECKING PREREQUISITES")
        print("="*50)

        required_modules = [
            "selenium", "pyautogui", "pyperclip", "pygetwindow", "tkinter",
            "pandas", "openpyxl", "tqdm", "webdriver_manager", "re"
        ]
        missing = []
        for mod in required_modules:
            if importlib.util.find_spec(mod) is None:
                missing.append(mod)
        if missing:
            print(f"❌ Missing packages: {', '.join(missing)}")
            print(f"Please run: pip install {' '.join(missing)}")
            messagebox.showerror("Missing Packages", f"Please install:\n" + "\n".join(missing))
            sys.exit(1)
        else:
            print("✅ All required Python packages are installed.")

        # Check Chrome debug status
        if not self._is_chrome_running_with_debug():
            print("⚠️ Chrome not running with debugging port")
            if not self._launch_chrome_with_debug():
                messagebox.showerror("Chrome Error", "Failed to launch Chrome with debugging")
                sys.exit(1)
            self._chrome_was_launched = True
        else:
            print("✅ Chrome is running with debugging port.")
        return True

    def _is_chrome_running_with_debug(self):
        """Check if Chrome is running with debug port and is a chrome.exe process.
        If port is open but not by chrome.exe, kill the process and exit."""
        print("[CHECK] Checking if Chrome is running with debug port...")
        try:
            result = subprocess.run(
                ['netstat', '-ano'],
                capture_output=True,
                text=True,
                creationflags=subprocess.CREATE_NO_WINDOW
            )
            lines = result.stdout.splitlines()
            for line in lines:
                if ":9222" in line and "LISTENING" in line:
                    parts = line.split()
                    pid = parts[-1]
                    # Now check if this PID is a Chrome process
                    tasklist = subprocess.run(
                        ['tasklist', '/FI', f'PID eq {pid}'],
                        capture_output=True,
                        text=True,
                        creationflags=subprocess.CREATE_NO_WINDOW
                    )
                    if "chrome.exe" in tasklist.stdout.lower():
                        print(f"[CHECK] Chrome debug port found and process is chrome.exe (PID {pid})")
                        return True
                    else:
                        print(f"[CHECK] Port 9222 is open but not used by chrome.exe (PID {pid})")
                        # Try to kill the process using the port
                        try:
                            subprocess.run(['taskkill', '/PID', pid, '/F'], creationflags=subprocess.CREATE_NO_WINDOW)
                            print(f"[CHECK] Killed process PID {pid} using port 9222.")
                        except Exception as kill_err:
                            print(f"[CHECK] Failed to kill process PID {pid}: {kill_err}")
                        print("[CHECK] Exiting script due to invalid debug port usage.")
                        import sys
                        sys.exit(1)
            print("[CHECK] Chrome debug port not found or not a Chrome process.")
            return False
        except Exception as e:
            print(f"[CHECK] Error checking Chrome debug port: {e}")
            return False

    def _launch_chrome_with_debug(self):
        print("[LAUNCH] Attempting to launch Chrome with debugging...")
        chrome_paths = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
        ]
        for path in chrome_paths:
            if os.path.exists(path):
                print(f"🚀 Launching Chrome with debugging: {path}")
                try:
                    subprocess.Popen([
                        path,
                        "--remote-debugging-port=9222",
                        "--user-data-dir=C:\\Temp\\ChromeDebugProfile",
                        "https://works.haryana.gov.in/HEWP_Login/login.aspx"
                    ], creationflags=subprocess.CREATE_NO_WINDOW)
                    time.sleep(0.2)
                    print("[LAUNCH] Chrome launched. Please login and rerun script.")
                    sys.exit(0)
                except Exception as e:
                    print(f"Failed to launch Chrome: {str(e)}")
                    return False

        print("❌ Chrome not found in standard locations")
        print("Please manually start Chrome with:")
        print("chrome.exe --remote-debugging-port=9222 --user-data-dir=C:\\Temp\\ChromeDebugProfile")
        return False

    def connect_to_browser(self):
        """Connect to existing Chrome browser instance"""
        print("[BROWSER] Connecting to Chrome browser...")
        chrome_options = Options()
        chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
        print("="*50)
        print("BROWSER CONNECTION")
        print("="*50)
        try:
            self.driver = webdriver.Chrome(options=chrome_options)
            self.wait = WebDriverWait(self.driver, 20)
            if self._chrome_was_launched:
                print("\n⚠️ NEW CHROME SESSION DETECTED")
                print("Please complete login to HEWP in the Chrome window")
                print("After login, rerun this script to continue automation.")
                sys.exit(0)
            else:
                print("✅ Reconnected to existing Chrome session")
            print("="*50)
            print("STARTING AUTOMATION")
            print("="*50)
        except Exception as e:
            print(f"❌ Browser connection failed: {str(e)}")
            sys.exit(1)

    def _verify_logged_in(self):
        """CHANGE: Verify login by checking lblusername text is non-empty"""
        try:
            user_elem = WebDriverWait(self.driver, 5).until(
                EC.presence_of_element_located((By.ID, "lblusername"))
            )
            username = user_elem.text.strip()
            print(f"[LOGIN] Username label: '{username}'")
            return bool(username)
        except Exception:
            return False
    def _switch_to_hewp_tab(self):
        """CHANGE: Switch to a tab whose URL contains works.haryana.gov.in"""
        for handle in self.driver.window_handles:  # [2][1][4]
            self.driver.switch_to.window(handle)
            url = self.driver.current_url
            print(f"[TAB] Checking tab: {url}")
            if "works.haryana.gov.in" in url:
                print("[TAB] Switched to HEWP tab.")
                return True
        return False  # CHANGE
    def _navigate_to_submit_bill(self):
        """CHANGE: Navigate left menu to Submit Bill to JE → Est_Add_Items_emb.aspx"""
        print("[NAV] Navigating to 'Submit Bill to JE' page ...")
        try:
            # Ensure window visible
            self.ensure_window_visible()

            # Sidebar anchor with href '#actmenucon' and text contains 'Submit Bill to JE'
            sidebar_links = self.driver.find_elements(
                By.XPATH,
                "//a[contains(@href, '#actmenucon') and contains(., 'Submit Bill to JE')]"
            )
            sidebar_link = next((l for l in sidebar_links if l.is_displayed()), None)
            if sidebar_link:
                self.driver.execute_script("arguments.scrollIntoView();", sidebar_link)
                try:
                    sidebar_link.click()
                except Exception:
                    self.driver.execute_script("arguments.click();", sidebar_link)
                print("[NAV] Sidebar 'Submit Bill to JE' clicked.")
                time.sleep(1)
            else:
                print("[NAV] Sidebar 'Submit Bill to JE' not found.")

            # Submenu link to Est_Add_Items_emb.aspx
            submenu_links = self.driver.find_elements(
                By.XPATH,
                "//ul[contains(@id, 'actmenucon')]/li/a[contains(@href, '/E-Billing/Est_Add_Items_emb.aspx') and contains(., 'Submit Bill to JE')]"
            )
            submenu_link = next((l for l in submenu_links if l.is_displayed()), None)
            if submenu_link:
                self.driver.execute_script("arguments.scrollIntoView();", submenu_link)
                try:
                    submenu_link.click()
                except Exception:
                    self.driver.execute_script("arguments.click();", submenu_link)
                print("[NAV] Submenu 'Submit Bill to JE' clicked.")
                time.sleep(2)
                return True  # CHANGE
            else:
                print("[NAV] Submenu 'Submit Bill to JE' link not found.")
                return False  # CHANGE
        except Exception as e:
            print(f"[NAV] Navigation error: {e}")
            return False  # CHANGE
    def ensure_on_target_page(self):
        """CHANGE: Gate: correct domain, logged in, and on the Submit Bill page with dropdowns present"""
        print("[GATE] Ensuring correct domain/tab, login, and page context...")
        current_url = ""
        try:
            current_url = self.driver.current_url
        except Exception:
            pass
        print(f"[GATE] Current URL: {current_url}")

        # 1) Domain/tab gate
        if "works.haryana.gov.in" not in current_url:
            print("[GATE] Not on works.haryana.gov.in, searching tabs...")
            if not self._switch_to_hewp_tab():  # [2][1][4]
                print("[GATE] No HEWP tab found; navigating to contractor home.")
                self.driver.get("https://works.haryana.gov.in/contractor/contractorHome.aspx")
                time.sleep(2)

        # 2) Login gate
        if not self._verify_logged_in():
            print("[GATE] Not logged in. Please login in Chrome and rerun.")
            messagebox.showerror("Login Required", "Please login to HEWP in Chrome, then rerun this tool.")
            sys.exit(0)  # CHANGE: exit cleanly to avoid silent idle

        # 3) Page identity gate: ensure dropdown page is open
        # If ddlcomp (or ddltender) not present, navigate via sidebar
        try:
            self.driver.find_element(By.ID, "ddlcomp")
            print("[GATE] Found ddlcomp.")
        except NoSuchElementException:
            print("[GATE] ddlcomp not found; trying to navigate to 'Submit Bill to JE' page.")
            if not self._navigate_to_submit_bill():
                messagebox.showerror("Navigation Error", "Could not open 'Submit Bill to JE' page. Please open it manually and rerun.")
                sys.exit(0)  # CHANGE
            # After navigation, wait for ddlcomp presence
            try:
                WebDriverWait(self.driver, 10).until(EC.presence_of_element_located((By.ID, "ddlcomp")))
                print("[GATE] ddlcomp present after navigation.")
            except TimeoutException:
                messagebox.showerror("Navigation Error", "Page did not load ddlcomp. Please try again.")
                sys.exit(0)  # CHANGE
        
    def ensure_window_visible(self):
        """Bring Chrome window to foreground if minimized"""
        print("[WINDOW] Ensuring Chrome window is visible...")
        try:
            chrome_windows = gw.getWindowsWithTitle("Chrome")
            if chrome_windows:
                chrome_window = chrome_windows[0]
                if chrome_window.isMinimized:
                    print("[WINDOW] Restoring minimized Chrome window.")
                    chrome_window.restore()
                chrome_window.activate()
                print("[WINDOW] Chrome window activated.")
                time.sleep(0.5)
            else:
                print("[WINDOW] No Chrome window found.")
        except Exception as e:
            print(f"Window management warning: {str(e)}")


    def get_dropdown_options(self, dropdown_id):
        """Get all non-empty, non-placeholder options from a dropdown"""
        try:
            dropdown = Select(self.wait.until(
                EC.presence_of_element_located((By.ID, dropdown_id))
            ))
            return [
                {'value': opt.get_attribute("value"), 'text': opt.text.strip()}
                for opt in dropdown.options
                if opt.get_attribute("value")
                and opt.text.strip()
                and opt.text.strip().lower() not in ["select one", "select", "choose", "--select--"]
            ]
        except Exception as e:
            print(f"⚠ Error reading {dropdown_id} options: {str(e)}")
            return []

    def safe_select_dropdown(self, dropdown_id, value):
        """Select dropdown option with retries and proper waiting"""
        for attempt in range(self.max_retries):
            try:
                dropdown = Select(self.wait.until(
                    EC.element_to_be_clickable((By.ID, dropdown_id))
                ))
                dropdown.select_by_value(value)
                
                # Wait for the dropdown value to update
                self.wait.until(lambda d: dropdown.first_selected_option.get_attribute("value") == value)
                # Wait for dependent element to update (if needed)
                self.wait.until(EC.presence_of_element_located((By.ID, "lbltobeexecuted")))
                return True
                
            except Exception as e:
                print(f"⚠ Attempt {attempt+1} failed for {dropdown_id}={value}: {str(e)}")
                time.sleep(0.5)  # Reduced sleep
        
        print(f"✖ Failed to select {dropdown_id}={value} after {self.max_retries} attempts")
        return False

    def extract_quantity(self):
        """Extract the quantity to be executed"""
        try:
            return self.driver.find_element(By.ID, "lbltobeexecuted").text.strip()
        except:
            return "N/A"

    def extract_table_data(self):
        """Extract all data from the main table"""
        try:
        
            try:
                savecheck_btn = self.wait.until(
                    EC.element_to_be_clickable((By.ID, "btn_Final_Save"))
                )
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", savecheck_btn)
            except TimeoutException:
                print("Page not loaded properly.")
                return []
            
            # Get DNIT unit from lbltobeexecutedunit
            try:
                dnit_unit = self.driver.find_element(By.ID, "lbltobeexecutedunit").text.strip()
            except Exception:
                dnit_unit = ""

            try:
                table = WebDriverWait(self.driver, 0.5).until(
                    EC.presence_of_element_located((By.ID, "GV_Add_to_List_POP"))
                )
            except TimeoutException:
                print("Table not found. Assuming all rows blanks.")
                # Return a row with only Total Quantity as per DNIT
                table_data = []
                table_data.append({
                    'Description': '',
                    'Number': '',
                    'Length': '',
                    'Breadth': '',
                    'Depth': '',
                    'Qty': '',
                    'Unit': dnit_unit,
                    'Rate_Type': '',
                    'Rate': '',
                    'Amount': '',
                    'Quantity Previously Executed': ''
                    
                })
                return table_data
            
            # Scroll to the table to ensure it's loaded
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", table)
            time.sleep(0.5)  # Allow time for table to render
            rows = table.find_elements(By.TAG_NAME, "tr")
            
            # Skip header and footer rows
            data_rows = [row for row in rows[1:-1] if "total-row" not in row.get_attribute("class")]
            
            table_data = []
            for row in data_rows:
                cells = row.find_elements(By.TAG_NAME, "td")
                if len(cells) >= 11:
                   
                    try:
                        qty = float(cells[10].text.strip().replace(',', ''))
                    except ValueError:
                        qty = cells[10].text.strip()
                    try:
                        rate = float(cells[3].text.strip().replace(',', ''))
                    except ValueError:
                        rate = cells[3].text.strip()
                    try:
                        number = float(cells[5].text.strip().replace(',', ''))
                    except ValueError:
                        number = cells[5].text.strip()
                    try:
                        length = float(cells[6].text.strip().replace(',', ''))
                    except ValueError:
                        length = cells[6].text.strip()
                    try:
                        breadth = float(cells[7].text.strip().replace(',', ''))
                    except ValueError:
                        breadth = cells[7].text.strip()
                    try:
                        depth = float(cells[8].text.strip().replace(',', ''))
                    except ValueError:
                        depth = cells[8].text.strip()
                    # Only calculate amount if both qty and rate are numbers
                    if isinstance(qty, (int, float)) and isinstance(rate, (int, float)):
                        amount = qty * rate
                    else:
                        amount = ""

                    table_data.append({
                        'Description': cells[1].text.strip(),
                        'Number': number,
                        'Length': length,
                        'Breadth': breadth,
                        'Depth': depth,
                        'Qty': qty,
                        'Unit': cells[4].text.strip(),
                        'Rate_Type': cells[2].text.strip(),
                        'Rate': rate,
                        'Amount': amount,
                        'Quantity Previously Executed': cells[9].text.strip()
                        
                    })
            return table_data
        except Exception as e:
            print(f"⚠ Error extracting table data: {str(e)}")
            return []

    def scrape_all_combinations(self):
        """Main scraping logic to process all dropdown combinations"""
        print("\nStarting data extraction process...")

        # Get all main head options, skip empty and "Select One"
        main_heads = [
            mh for mh in self.get_dropdown_options("ddlcomp")
            if mh['text'].strip() and mh['text'].strip().lower() != "select one"
        ]
        print(f"Found {len(main_heads)} Main Head options (excluding empty/'Select One')")

        # Select main head
        if not main_heads:
            print("No Main Head options found.")
            return

        if len(main_heads) == 1:
            selected_main_heads = [main_heads[0]]
            print(f"Only one Main Head found: {main_heads[0]['text']}. Selecting automatically.")
        else:
            selected = select_option(main_heads, "Select Main Head", "Multiple Main Heads found. Please select one:")
            selected_main_heads = [selected]

        for main_head in tqdm(selected_main_heads, desc="Main Heads"):
            print(f"\nProcessing Main Head: {main_head['text']}")

            if not self.safe_select_dropdown("ddlcomp", main_head['value']):
                continue

            # Get sub heads for current main head, skip empty and "Select One"
            sub_heads = [
                sh for sh in self.get_dropdown_options("ddlsubhead")
                if sh['text'].strip() and sh['text'].strip().lower() != "select one"
            ]
            print(f"├─ Found {len(sub_heads)} Sub Head options (excluding empty/'Select One')")

            if not sub_heads:
                print("No Sub Head options found.")
                continue

            if len(sub_heads) == 1:
                selected_sub_heads = [sub_heads[0]]
                print(f"Only one Sub Head found: {sub_heads[0]['text']}. Selecting automatically.")
            else:
                selected = select_option(sub_heads, "Select Sub Head", "Multiple Sub Heads found. Please select one:")
                selected_sub_heads = [selected]

            for sub_head in selected_sub_heads:
                if not self.safe_select_dropdown("ddlsubhead", sub_head['value']):
                    continue

                # Get items for current sub head, skip empty and "Select One"
                items = [
                    it for it in self.get_dropdown_options("ddlitemnumber")
                    if it['text'].strip() and it['text'].strip().lower() != "select one"
                ]

                for item in items:
                    if not self.safe_select_dropdown("ddlitemnumber", item['value']):
                        continue

                    print(f"├─ Processing Sub Head: {sub_head['text']} - Item: {item['text']}") 
                    # Extract data for this combination
                    quantity = self.extract_quantity()
                    table_data = self.extract_table_data()

                    # Combine all data with dropdown info
                    for row in table_data:
                        # Extract only the text inside [ ]
                        match = re.search(r'\[([^\]]+)\]', item['text'])
                        hsr_item_number = match.group(1).strip() if match else item['text'].strip()


                        # Try to convert quantity to float, else keep as text
                        try:
                            total_quantity = float(quantity.replace(',', ''))
                        except (ValueError, AttributeError):
                            total_quantity = quantity

                        self.all_data.append({
                            'Main_Head': main_head['text'],
                            'Sub_Head': sub_head['text'],
                            'HSR Item_Number': hsr_item_number,
                            'Total Quantity as per DNIT': total_quantity,
                            **row
                        })

        print(f"\n✔ Completed! Extracted {len(self.all_data)} total records")

    def save_to_excel(self, filename="haryana_ebilling_data.xlsx", selected_sub_head=None):
        """Save all collected data to Excel file with permission handling and summary sheet"""
        if not self.all_data:
            print("No data to save")
            return None

        df = pd.DataFrame(self.all_data)

        # Reorder columns for better organization
        dropdown_cols = [
            'Main_Head',
            'Sub_Head',
            'HSR Item_Number',
            'Total Quantity as per DNIT'
        ]
        other_cols = [col for col in df.columns if col not in dropdown_cols]

        max_attempts = 3
        attempt = 0
        success = False
        save_path = filename

        while attempt < max_attempts and not success:
            attempt += 1
            try:
                if attempt > 1 and os.path.exists(filename):
                    # Create a backup filename if primary fails
                    timestamp = time.strftime("%Y%m%d_%H%M%S")
                    desktop_dir = os.path.join(os.path.expanduser('~'), 'Desktop')
                    save_path = os.path.join(desktop_dir, f"haryana_ebilling_data_{timestamp}.xlsx")
                    

                with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                    # Always write main data
                    df[dropdown_cols + other_cols].to_excel(writer, index=False, sheet_name="Sheet1")

                    try:
                        # --- SUMMARY SHEET LOGIC STARTS HERE ---
                        summary_cols = [
                            "HSR Item_Number", "Qty", "Unit", "Rate", "Amount", "Total Quantity as per DNIT"
                        ]
                        df_summary = df.copy()
                        df_summary["Qty"] = pd.to_numeric(df_summary["Qty"], errors="coerce").fillna(0)
                        df_summary["Amount"] = pd.to_numeric(df_summary["Amount"], errors="coerce").fillna(0)
                        df_summary["Total Quantity as per DNIT"] = pd.to_numeric(df_summary["Total Quantity as per DNIT"], errors="coerce").fillna(0)

                        summary = (
                            df_summary.groupby("HSR Item_Number", as_index=False)
                            .agg({
                                "Qty": "sum",
                                "Unit": "first",
                                "Rate": "first",
                                "Amount": "sum",
                                "Total Quantity as per DNIT": "sum"
                            })
                        )

                        summary.insert(0, "Sr. No.", range(1, len(summary) + 1))

                        # Create the summary sheet
                        workbook = writer.book
                        summary_sheet = workbook.create_sheet("Portal Summary")

                        # Add selected sub head to the summary title
                        summary_title = "Data from: Portal"
                        if selected_sub_head:
                            summary_title += f" ({selected_sub_head})"
                        summary_sheet.append([summary_title])
                        summary_sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
                        cell = summary_sheet.cell(row=1, column=1)
                        cell.font = XLFont(bold=True)
                        cell.alignment = Alignment(horizontal="center")

                        summary_headers = ["Sr. No.", "HSR Item_Number", "Qty", "Unit", "Rate", "Amount", "Total Quantity as per DNIT"]
                        summary_sheet.append(summary_headers)

                        for row in summary.itertuples(index=False):
                            summary_sheet.append(row)

                        total_qty = summary["Qty"].sum()
                        total_amount = summary["Amount"].sum()
                        total_dnit = summary["Total Quantity as per DNIT"].sum()
                        total_row = ["Total", "", total_qty, "", "", total_amount, total_dnit]
                        summary_sheet.append(total_row)

                        for col in range(1, 6):
                            cell = summary_sheet.cell(row=summary_sheet.max_row, column=col)
                            cell.font = XLFont(bold=True)

                        # Auto-adjust column widths
                        for ws in [writer.sheets['Sheet1'], summary_sheet]:
                            for column_cells in ws.columns:
                                for cell in column_cells:
                                    if not isinstance(cell, openpyxl.cell.cell.MergedCell):
                                        col_letter = cell.column_letter
                                        break
                                else:
                                    continue
                                max_length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=0)
                                ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

                    except Exception as summary_e:
                        print(f"⚠ Error creating summary or formatting: {summary_e}")
                        print("Main data sheet will still be saved.")

                print(f"✔ Data successfully saved to {save_path}")
                success = True
                return save_path

            except PermissionError as e:
                print(f"⚠ Attempt {attempt} failed: {str(e)}")
                if attempt < max_attempts:
                    if os.path.exists(filename):
                        print(f"File '{filename}' may be locked. Please close it if open in Excel.")
                    print("Retrying...")
                    time.sleep(2)  # Wait before retrying
                else:
                    # Try saving to Documents folder as last resort
                    documents_path = os.path.join(os.path.expanduser('~'), 'Documents')
                    if os.path.exists(documents_path):
                        fallback_path = os.path.join(documents_path, f"haryana_ebilling_data_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
                        try:
                            df[dropdown_cols + other_cols].to_excel(fallback_path, index=False)
                            print(f"✔ Data saved to fallback location: {fallback_path}")
                            success = True
                            return fallback_path  # Return fallback path
                        except Exception as fallback_e:
                            print(f"✖ Failed to save to fallback location: {str(fallback_e)}")
                            print("Please check your permissions or try a different directory.")
            except Exception as e:
                print(f"✖ Unexpected error saving Excel file: {str(e)}")
                break
        return None  # If all attempts fail

    def close(self):
        if self.driver:
            # self.driver.quit()
            print("skipped Closing browser session...")
            pass

def select_option(options, title, prompt):
    """Show a GUI selection window for options with enhanced UI/UX. Returns the selected option dict or None if cancelled."""
    if len(options) == 1:
        return options[0]

    selected = {}

    filtered_indices = list(range(len(options)))

    def on_ok(event=None):
        selection = listbox.curselection()
        if selection:
            idx = filtered_indices[selection[0]]
            selected['option'] = [options[idx] for idx in selection] if multi_select else options[idx]
            root.destroy()

    def on_cancel(event=None):
        selected['option'] = None if not multi_select else []
        root.destroy()

    def on_search(*args):
        search_term = search_var.get().lower()
        listbox.delete(0, tk.END)
        filtered_indices.clear()
        for i, opt in enumerate(options):
            if search_term in opt['text'].lower():
                display_text = f"{i+1}. {opt['text']}"
                listbox.insert(tk.END, display_text)
                filtered_indices.append(i)
        # Restore selection if possible
        if listbox.size() > 0:
            listbox.selection_set(0)

    def select_all(event=None):
        listbox.selection_set(0, tk.END)
        return "break"  # Prevent default behavior

    def show_tooltip(event):
        if tooltip_time_id:
            root.after_cancel(tooltip_time_id)
        widget = event.widget
        index = widget.nearest(event.y)
        if index >= 0:
            bbox = widget.bbox(index)
            if bbox and (event.x > bbox[0] and event.x < bbox[0] + bbox[2]):
                text = widget.get(index)
                if len(text) > 50:  # Only show tooltip for long texts
                    def delayed_show():
                        tooltip.config(text=text)
                        tooltip.place(x=event.x_root - root.winfo_rootx() + 15,
                                      y=event.y_root - root.winfo_rooty() + 15)
                    tooltip_time_id = root.after(500, delayed_show)

    def hide_tooltip(event):
        if tooltip_time_id:
            root.after_cancel(tooltip_time_id)
        tooltip.place_forget()

    def show_context_menu(event):
        try:
            index = listbox.nearest(event.y)
            if index >= 0:
                listbox.selection_clear(0, tk.END)
                listbox.selection_set(index)
                context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()

    def copy_selected():
        selection = listbox.curselection()
        if selection:
            root.clipboard_clear()
            root.clipboard_append(options[selection[0]]['text'])

    # Configuration
    multi_select = False  # Set to True if you want multi-selection
    min_width, min_height = 600, 300

    root = tk.Tk()
    root.title(title)
    max_height = root.winfo_screenheight() - 100

    # Calculate dynamic height
    item_count = len(options)
    dynamic_height = min(max(min_height, item_count * 20 + 150), max_height)

    root.geometry(f"{min_width}x{dynamic_height}")
    root.minsize(min_width, min(min_height, dynamic_height))
    root.attributes("-topmost", True)

    # Setup styles
    style = ttk.Style()
    style.theme_use('clam')

    # Custom fonts
    title_font = Font(family="Segoe UI", size=11, weight="bold")
    text_font = Font(family="Segoe UI", size=10)
    button_font = Font(family="Segoe UI", size=10, weight="bold")

    style.configure('TButton', font=button_font, padding=5)
    style.configure('Accent.TButton', background='#0078d7', foreground='white')
    style.map('Accent.TButton',
              background=[('active', '#005499'), ('pressed', '#003366')])
    style.configure('TEntry', font=text_font, padding=5)
    style.configure('TLabel', font=title_font)

    # Main frame
    main_frame = ttk.Frame(root, padding=15)
    main_frame.grid(row=0, column=0, sticky="nsew")
    root.rowconfigure(0, weight=1)
    root.columnconfigure(0, weight=1)

    # Prompt label
    ttk.Label(main_frame, text=prompt).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 10))

    # Search box
    search_var = tk.StringVar()
    search_box = ttk.Entry(main_frame, textvariable=search_var)
    search_box.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 10))
    search_var.trace_add('write', on_search)
    search_box.focus_set()

    # Listbox with scrollbar
    list_frame = ttk.Frame(main_frame)
    list_frame.grid(row=2, column=0, columnspan=2, sticky="nsew")
    main_frame.rowconfigure(2, weight=1)
    main_frame.columnconfigure(0, weight=1)

    scrollbar = ttk.Scrollbar(list_frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    listbox = tk.Listbox(
        list_frame,
        font=text_font,
        selectmode=tk.MULTIPLE if multi_select else tk.SINGLE,
        yscrollcommand=scrollbar.set,
        bg='white',
        selectbackground='#0078d7',
        selectforeground='white',
        activestyle='none',
        relief='flat',
        highlightthickness=0
    )
    listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.config(command=listbox.yview)

    # Populate list
    for i, opt in enumerate(options):
        listbox.insert(tk.END, f"{i+1}. {opt['text']}")
    if listbox.size() > 0:
        listbox.selection_set(0)

    # Button frame (always at the bottom)
    btn_frame = ttk.Frame(main_frame)
    btn_frame.grid(row=9, column=0, columnspan=2, sticky="ew", pady=(5, 2))

    ok_btn = ttk.Button(
        btn_frame,
        text="OK",
        command=on_ok,
        style='Accent.TButton'
    )
    ok_btn.pack(side=tk.LEFT, padx=(10, 5), ipadx=2, ipady=2)

    cancel_btn = ttk.Button(
        btn_frame,
        text="Cancel",
        command=on_cancel
    )
    cancel_btn.pack(side=tk.RIGHT, padx=(5, 10), ipadx=2, ipady=2)

    # Tooltip
    tooltip = ttk.Label(root, background="#ffffe0", relief=tk.SOLID, borderwidth=1, font=text_font)
    tooltip_time_id = None

    # Context menu
    context_menu = tk.Menu(root, tearoff=0, font=text_font)
    context_menu.add_command(label="Copy Text", command=copy_selected)

    # Bindings
    root.bind('<Return>', on_ok)
    root.bind('<Escape>', on_cancel)
    listbox.bind('<Double-1>', on_ok)
    listbox.bind('<Control-a>', select_all)
    listbox.bind('<Motion>', show_tooltip)
    listbox.bind('<Leave>', hide_tooltip)
    listbox.bind('<Button-3>', show_context_menu)
    search_box.bind('<Control-a>', lambda e: search_box.selection_range(0, tk.END))

    # Focus search box by default
    search_box.focus_set()

    root.protocol("WM_DELETE_WINDOW", on_cancel)
    root.mainloop()

    return selected.get('option', None if not multi_select else [])



    

# In your main function:
def main():
    print("Haryana E-Billing Data Extractor")
    print("=" * 50)
    
    # Standard Desktop path
    desktop_dir = os.path.join(os.path.expanduser('~'), 'Desktop')
    # OneDrive Desktop path (common for Windows)
    onedrive_dir = os.path.join(os.path.expanduser('~'), 'OneDrive', 'Desktop')

    # Prefer standard Desktop, else OneDrive Desktop, else create standard Desktop
    if os.path.exists(desktop_dir):
        chosen_dir = desktop_dir
    elif os.path.exists(onedrive_dir):
        chosen_dir = onedrive_dir
    else:
        os.makedirs(desktop_dir, exist_ok=True)
        chosen_dir = desktop_dir

    timestamp = time.strftime("%Y%m%d_%H%M%S")
    default_filename = os.path.join(chosen_dir, f"haryana_ebilling_data_{timestamp}.xlsx")
    
    filename = default_filename
    print(f"Using Desktop for output file: {filename}")
    
    scraper = HaryanaEBillingScraper(
        "https://works.haryana.gov.in/E-Billing/Est_Add_Items_emb.aspx#"
    )
    
    try:
        print("[START] Starting the scraping process...")
            
        scraper.ensure_window_visible()
        scraper.ensure_on_target_page()
        scraper.scrape_all_combinations()
        
        # Save and get the actual saved path
        saved_path = scraper.save_to_excel(filename)
        
        # After scraping, get the selected sub head (example: from the last scraped row)
        selected_sub_head = scraper.all_data[-1]['Sub_Head'] if scraper.all_data else None
        saved_path = scraper.save_to_excel(filename, selected_sub_head=selected_sub_head)
        
        # Auto-open the Excel file only if it exists
        if saved_path and os.path.exists(saved_path):
            try:
                os.startfile(saved_path)
                print(f"Opened Excel file: {saved_path}")
                sys.exit(0)  # Exit gracefully, no error
            except Exception as e:
                print(f"Could not open Excel file automatically: {e}")
        else:
            print("Excel file was not created (no data extracted).")
        
    except Exception as e:
        print(f"\n✖ Critical error: {str(e)}")
        
    finally:
        scraper.close()
if __name__ == "__main__":
    main()
