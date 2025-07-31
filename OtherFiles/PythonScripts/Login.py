import os
import sys
import ctypes
import time
import subprocess
import importlib.util

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pyautogui
import pyperclip
import tkinter as tk
from tkinter import Toplevel, Label, messagebox, simpledialog
from tkinter import ttk
import pygetwindow as gw
import pytesseract
import openpyxl
from openpyxl import Workbook
from PIL import ImageEnhance, ImageFilter
import cv2
import numpy as np

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

def lighten_color(hex_color, factor=0.2):
    hex_color = hex_color.lstrip('#')
    rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    lighter = tuple(min(255, int(c + (255 - c) * factor)) for c in rgb)
    return f'#{lighter[0]:02x}{lighter[1]:02x}{lighter[2]:02x}'

def darken_color(hex_color, factor=0.2):
    hex_color = hex_color.lstrip('#')
    rgb = tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    darker = tuple(max(0, int(c * (1 - factor))) for c in rgb)
    return f'#{darker[0]:02x}{darker[1]:02x}{darker[2]:02x}'

class RoundedFrame(tk.Canvas):
    def __init__(self, master=None, radius=25, color="#ffffff", width=200, height=200, *args, **kwargs):
        super().__init__(master, width=width, height=height, *args, **kwargs)
        self.radius = radius
        self.color = color
        self.bind("<Configure>", self.draw)
    def draw(self, event=None):
        self.delete("all")
        width = self.winfo_width()
        height = self.winfo_height()
        self.create_round_rect(3, 3, width+3, height+3, radius=self.radius, fill="#e0e0e0", outline="#e0e0e0")
        self.create_round_rect(0, 0, width, height, radius=self.radius, fill=self.color, outline=self.color)
        self.tag_lower("all")
    def create_round_rect(self, x1, y1, x2, y2, radius=25, **kwargs):
        points = [x1+radius, y1, x1+radius, y1, x2-radius, y1, x2-radius, y1, x2, y1,
                  x2, y1+radius, x2, y1+radius, x2, y2-radius, x2, y2-radius, x2, y2,
                  x2-radius, y2, x2-radius, y2, x1+radius, y2, x1+radius, y2, x1, y2,
                  x1, y2-radius, x1, y2-radius, x1, y1+radius, x1, y1+radius, x1, y1]
        return self.create_polygon(points, **kwargs, smooth=True)

class RoundedButton(tk.Canvas):
    def __init__(self, master=None, text="", radius=25, btnforeground="#000000", btnbackground="#ffffff", 
                 clicked=None, width=None, height=None, hover_color=None, active_color=None, *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.btnbackground = btnbackground
        self.btnforeground = btnforeground
        self.hover_color = hover_color if hover_color else lighten_color(btnbackground, 0.2)
        self.active_color = active_color if active_color else darken_color(btnbackground, 0.2)
        self.clicked = clicked
        self.radius = radius
        self.width = width if width else 120
        self.height = height if height else 40
        self.configure(width=self.width, height=self.height, bd=0, highlightthickness=0, bg=master.cget("bg"))
        self.text = self.create_text(self.width//2, self.height//2, text=text, fill=btnforeground, font=("Segoe UI", 10, "bold"))
        self.bind("<Enter>", self.on_enter)
        self.bind("<Leave>", self.on_leave)
        self.bind("<ButtonPress-1>", self.on_press)
        self.bind("<ButtonRelease-1>", self.on_release)
        self.draw_button()
    def draw_button(self, color=None):
        self.delete("button")
        color = color if color else self.btnbackground
        self.create_round_rect(2, 2, self.width+2, self.height+2, radius=self.radius, fill="#cccccc", outline="#cccccc", tags="shadow")
        self.tag_lower("shadow")
        self.create_round_rect(0, 0, self.width, self.height, radius=self.radius, fill=color, outline=color, tags="button")
        self.tag_raise(self.text)
    def create_round_rect(self, x1, y1, x2, y2, radius=25, **kwargs):
        points = [x1+radius, y1, x1+radius, y1, x2-radius, y1, x2-radius, y1, x2, y1,
                  x2, y1+radius, x2, y1+radius, x2, y2-radius, x2, y2-radius, x2, y2,
                  x2-radius, y2, x2-radius, y2, x1+radius, y2, x1+radius, y2, x1, y2,
                  x1, y2-radius, x1, y2-radius, x1, y1+radius, x1, y1+radius, x1, y1]
        return self.create_polygon(points, **kwargs, smooth=True)
    def on_enter(self, event):
        self.draw_button(self.hover_color)
        self.config(cursor="hand2")
        self.scale("all", self.width/2, self.height/2, 1.05, 1.05)
    def on_leave(self, event):
        self.draw_button(self.btnbackground)
        self.scale("all", self.width/2, self.height/2, 1/1.05, 1/1.05)
    def on_press(self, event):
        self.draw_button(self.active_color)
        self.scale("all", self.width/2, self.height/2, 0.95, 0.95)
    def on_release(self, event):
        self.draw_button(self.hover_color)
        self.scale("all", self.width/2, self.height/2, 1.05/0.95, 1.05/0.95)
        if self.clicked:
            self.clicked()

class RoundedEntry(tk.Frame):
    def __init__(self, master=None, radius=10, *args, **kwargs):
        super().__init__(master, bg=master.cget("bg"))
        self.radius = radius
        self.canvas = tk.Canvas(self, bd=0, highlightthickness=0, bg=master.cget("bg"), height=48)
        self.canvas.pack(fill="both", expand=True)
        self.entry = tk.Entry(self, *args, **kwargs)
        self.entry.place(x=10, y=12, relwidth=0.96, height=28)
        self.entry.bind("<FocusIn>", self.on_focus_in)
        self.entry.bind("<FocusOut>", self.on_focus_out)
        self.bind("<Configure>", lambda e: self.redraw())
        self.redraw()

        # Clipboard bindings for copy, cut, paste
        self.entry.bind("<Control-c>", self._copy)
        self.entry.bind("<Control-x>", self._cut)
        self.entry.bind("<Control-v>", self._paste)
        self.entry.bind("<Button-3>", self._show_context_menu)  # Right-click

    def _copy(self, event=None):
        self.entry.event_generate("<<Copy>>")
        return "break"

    def _cut(self, event=None):
        self.entry.event_generate("<<Cut>>")
        return "break"

    def _paste(self, event=None):
        self.entry.event_generate("<<Paste>>")
        return "break"

    def _show_context_menu(self, event):
        menu = tk.Menu(self.entry, tearoff=0)
        menu.add_command(label="Cut", command=lambda: self.entry.event_generate("<<Cut>>"))
        menu.add_command(label="Copy", command=lambda: self.entry.event_generate("<<Copy>>"))
        menu.add_command(label="Paste", command=lambda: self.entry.event_generate("<<Paste>>"))
        menu.tk_popup(event.x_root, event.y_root)

    def redraw(self):
        self.canvas.delete("all")
        width = self.winfo_width()
        height = 48
        self.create_round_rect(2, 2, width-2, height-2, radius=self.radius, fill="#e0e0e0", outline="#e0e0e0")
        self.create_round_rect(0, 0, width-4, height-4, radius=self.radius, fill="#ffffff", outline="#cccccc")

    def on_focus_in(self, event):
        self.canvas.delete("all")
        width = self.winfo_width()
        height = 48
        self.create_round_rect(2, 2, width-2, height-2, radius=self.radius, fill="#e0e0e0", outline="#e0e0e0")
        self.create_round_rect(0, 0, width-4, height-4, radius=self.radius, fill="#ffffff", outline="#4a90e2", width=2)

    def on_focus_out(self, event):
        self.redraw()

    def create_round_rect(self, x1, y1, x2, y2, radius=10, **kwargs):
        points = [x1+radius, y1, x1+radius, y1, x2-radius, y1, x2-radius, y1, x2, y1,
                  x2, y1+radius, x2, y1+radius, x2, y2-radius, x2, y2-radius, x2, y2,
                  x2-radius, y2, x2-radius, y2, x1+radius, y2, x1+radius, y2, x1, y2,
                  x1, y2-radius, x1, y2-radius, x1, y1+radius, x1, y1+radius, x1, y1]
        return self.canvas.create_polygon(points, **kwargs, smooth=True)

    # Proxy methods for Entry widget
    def get(self):
        return self.entry.get()
    def insert(self, index, value):
        self.entry.insert(index, value)
    def delete(self, first, last=None):
        self.entry.delete(first, last)
    def bind(self, sequence=None, func=None, add=None):
        self.entry.bind(sequence, func, add)

class HEWPLogin:
    def __init__(self):
        print("[INIT] Initializing HEWPLogin...")
        self.excel_file_path = r"C:\MRGARGSIR\OtherFiles\users\id.xlsx"
        self.notification_root = None
        self._chrome_was_launched = False

        self.title_font = ("Segoe UI", 20, "bold")
        self.subtitle_font = ("Segoe UI", 14)
        self.button_font = ("Segoe UI", 14, "bold")

        self._hide_console()

        if self._check_prerequisites():
            print("[INIT] Prerequisites OK. Connecting to browser...")
            self.connect_to_browser()
            self._hide_console()

    def _show_console(self):
        """Show console window"""
        if sys.platform == 'win32':
            print("[CONSOLE] Showing console window.")
            ctypes.windll.user32.ShowWindow(ctypes.windll.kernel32.GetConsoleWindow(), 1)

    def _hide_console(self):
        """Hide console window"""
        if sys.platform == 'win32':
            print("[CONSOLE] Hiding console window.")
            ctypes.windll.user32.ShowWindow(ctypes.windll.kernel32.GetConsoleWindow(), 0)

    def _check_prerequisites(self):
        """Verify all requirements are met"""
        print("[CHECK] Checking prerequisites...")
        print("="*50)
        print("CHECKING PREREQUISITES")
        print("="*50)

        required_modules = [
            "selenium", "pyautogui", "pyperclip", "pygetwindow",
            "pytesseract", "openpyxl",  "tkinter", "pandas", "tqdm", "webdriver_manager", "re"
        ]
        missing = []
        for mod in required_modules:
            if importlib.util.find_spec(mod) is None:
                missing.append(mod)
        if missing:
            print(f"❌ Missing packages: {', '.join(missing)}")
            print(f"Please run: pip install {' '.join(missing)}")
            messagebox.showerror("Missing Packages", f"Please install:\n" + "\n".join(missing))
            sys.exit(1)
        else:
            print("✅ All required Python packages are installed.")

        # Check Tesseract OCR executable
        tesseract_path = pytesseract.pytesseract.tesseract_cmd
        if not os.path.exists(tesseract_path):
            print(f"❌ Tesseract executable not found at {tesseract_path}")
            messagebox.showerror("Tesseract Missing", f"Tesseract not found at:\n{tesseract_path}\nPlease install Tesseract-OCR.")
            sys.exit(1)
        else:
            print(f"✅ Tesseract found at {tesseract_path}")

        # Check Excel file directory
        excel_dir = os.path.dirname(self.excel_file_path)
        if not os.path.exists(excel_dir):
            try:
                os.makedirs(excel_dir)
                print(f"✅ Created Excel directory: {excel_dir}")
            except Exception as e:
                print(f"❌ Could not create Excel directory: {e}")
                messagebox.showerror("Excel Directory Error", f"Could not create Excel directory:\n{excel_dir}")
                sys.exit(1)
        else:
            print(f"✅ Excel directory exists: {excel_dir}")

        # Check Chrome debug status
        if not self._is_chrome_running_with_debug():
            print("⚠️ Chrome not running with debugging port")
            if not self._launch_chrome_with_debug():
                messagebox.showerror("Chrome Error", "Failed to launch Chrome with debugging")
                sys.exit(1)
            self._chrome_was_launched = True
        else:
            print("✅ Chrome is running with debugging port.")
        return True  # Return True if checks passed

    def _is_chrome_running_with_debug(self):
        """Check if Chrome is running with debug port and is a chrome.exe process.
        If port is open but not by chrome.exe, kill the process, close driver, and exit."""
        print("[CHECK] Checking if Chrome is running with debug port...")
        try:
            result = subprocess.run(
                ['netstat', '-ano'],
                capture_output=True,
                text=True,
                creationflags=subprocess.CREATE_NO_WINDOW
            )
            lines = result.stdout.splitlines()
            for line in lines:
                if ":9222" in line and "LISTENING" in line:
                    parts = line.split()
                    pid = parts[-1]
                    # Now check if this PID is a Chrome process
                    tasklist = subprocess.run(
                        ['tasklist', '/FI', f'PID eq {pid}'],
                        capture_output=True,
                        text=True,
                        creationflags=subprocess.CREATE_NO_WINDOW
                    )
                    if "chrome.exe" in tasklist.stdout.lower():
                        print(f"[CHECK] Chrome debug port found and process is chrome.exe (PID {pid})")
                        return True
                    else:
                        print(f"[CHECK] Port 9222 is open but not used by chrome.exe (PID {pid})")
                        # Try to kill the process using the port
                        try:
                            subprocess.run(['taskkill', '/PID', pid, '/F'], creationflags=subprocess.CREATE_NO_WINDOW)
                            print(f"[CHECK] Killed process PID {pid} using port 9222.")
                        except Exception as kill_err:
                            print(f"[CHECK] Failed to kill process PID {pid}: {kill_err}")
                        # Close webdriver if open
                        if hasattr(self, 'driver') and self.driver:
                            try:
                                self.driver.quit()
                                print("[CHECK] Closed WebDriver.")
                            except Exception as e:
                                print(f"[CHECK] Error closing WebDriver: {e}")
                        print("[CHECK] Exiting script due to invalid debug port usage.")
                        sys.exit(1)
            print("[CHECK] Chrome debug port not found or not a Chrome process.")
            return False
        except Exception as e:
            print(f"[CHECK] Error checking Chrome debug port: {e}")
            return False

    def _launch_chrome_with_debug(self):
        """Launch Chrome with remote debugging"""
        print("[LAUNCH] Attempting to launch Chrome with debugging...")
        chrome_paths = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
        ]
        for path in chrome_paths:
            if os.path.exists(path):
                print(f"🚀 Launching Chrome with debugging: {path}")
                try:
                    subprocess.Popen([
                        path,
                        "--remote-debugging-port=9222",
                        "--user-data-dir=C:\\Temp\\ChromeDebugProfile",
                        "https://works.haryana.gov.in/HEWP_Login/login.aspx"
                    ], creationflags=subprocess.CREATE_NO_WINDOW)
                    time.sleep(0.5)
                    print("✅ Chrome launched successfully")
                    
                    return True
                except Exception as e:
                    print(f"Failed to launch Chrome: {str(e)}")
                    return False

        print("❌ Chrome not found in standard locations")
        print("Please manually start Chrome with:")
        print("chrome.exe --remote-debugging-port=9222 --user-data-dir=C:\\Temp\\ChromeDebugProfile")
        return False
        sys.exit(1)

    def connect_to_browser(self):
        """Connect to existing Chrome browser instance"""
        print("[BROWSER] Connecting to Chrome browser...")
        chrome_options = Options()
        chrome_options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
        print("="*50)
        print("BROWSER CONNECTION")
        print("="*50)
        try:
            self.driver = webdriver.Chrome(options=chrome_options)
            self.wait = WebDriverWait(self.driver, 20)
            if self._chrome_was_launched:
                print("Auto login to HEWP in the Chrome window")

            else:
                print("✅ Reconnected to existing Chrome session")
            print("="*50)
            print("STARTING AUTOMATION")
            print("="*50)
        except Exception as e:
            print(f"❌ Browser connection failed: {str(e)}")
            sys.exit(1)

    def ensure_window_visible(self):
        """Bring Chrome window to foreground if minimized"""
        print("[WINDOW] Ensuring Chrome window is visible...")
        try:
            chrome_windows = gw.getWindowsWithTitle("Chrome")
            if chrome_windows:
                chrome_window = chrome_windows[0]
                if chrome_window.isMinimized:
                    print("[WINDOW] Restoring minimized Chrome window.")
                    chrome_window.restore()
                chrome_window.activate()
                print("[WINDOW] Chrome window activated.")
                time.sleep(0.5)
            else:
                print("[WINDOW] No Chrome window found.")
        except Exception as e:
            print(f"Window management warning: {str(e)}")

    def show_notification(self, message, duration=3000):
        """Show a toast-style notification that actually works"""
        print(f"[NOTIFY] Showing notification: {message}")
        # Destroy any existing notification
        if self.notification_root:
            try:
                if self.notification_root.winfo_exists():
                    self.notification_root.destroy()
            except:
                pass

        # Create new notification
        self.notification_root = tk.Tk()
        self.notification_root.withdraw()  # Hide main window

        # Create popup window
        popup = Toplevel(self.notification_root)
        popup.wm_overrideredirect(True)

        # Calculate position (top-right corner)
        x_pos = self.notification_root.winfo_screenwidth() - 320
        popup.geometry(f"300x60+{x_pos}+50")

        # Style the popup
        popup.configure(background='#4CAF50')  # Green background
        Label(
            popup,
            text=message,
            fg='white',
            bg='#4CAF50',
            font=('Arial', 10, 'bold'),
            padx=20,
            pady=20
        ).pack()

        # Auto-close after duration
        def close_notification():
            try:
                if popup.winfo_exists():
                    popup.destroy()
                if self.notification_root and self.notification_root.winfo_exists():
                    self.notification_root.destroy()
                self.notification_root = None
            except:
                pass

        popup.after(duration, close_notification)

        # Force the window to update and show
        popup.update_idletasks()
        popup.deiconify()
        popup.lift()

        # Start the event loop (this will block until the notification is closed)
        self.notification_root.mainloop()

    def _ensure_excel_file(self):
        """Ensure the Excel file and directory exist, create if not."""
        dir_path = os.path.dirname(self.excel_file_path)
        if not os.path.exists(dir_path):
            os.makedirs(dir_path)
        if not os.path.exists(self.excel_file_path):
            wb = Workbook()
            ws = wb.active
            ws.append(["Name", "Username", "Password", "Description"])
            wb.save(self.excel_file_path)

    def _load_users_from_excel(self):
        """Load users from Excel file."""
        self._ensure_excel_file()
        wb = openpyxl.load_workbook(self.excel_file_path)
        ws = wb.active
        users = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(row):
                users.append(row)
        return users

    def _add_user_to_excel(self, name, username, password, description):
        """Add a new user to the Excel file."""
        self._ensure_excel_file()
        wb = openpyxl.load_workbook(self.excel_file_path)
        ws = wb.active
        ws.append([name, username, password, description])
        wb.save(self.excel_file_path)

    def user_details(self):
        """Premium user management dialog with animation and effects."""
        root = tk.Tk()
        root.withdraw()

        dialog = tk.Toplevel(root)
        dialog.withdraw()  # Start hidden

        dialog.title("🔐 MRGARGSIR Auto System")
        dialog.geometry("800x540")
        dialog.configure(bg="#f5f7fa")
        dialog.resizable(False, False)
        dialog.attributes("-topmost", True)
        dialog.update_idletasks()
        w, h = 800, 540
        x = (dialog.winfo_screenwidth() // 2) - (w // 2)
        y = (dialog.winfo_screenheight() // 2) - (h // 2)
        dialog.geometry(f"{w}x{h}+{x}+{y}")

        # Set custom window icon
        try:
            dialog.iconbitmap("security_icon.ico")
        except:
            pass

        # Main rounded frame
        main_frame = RoundedFrame(dialog, radius=30, color="#ffffff", width=780, height=520)
        main_frame.pack(padx=10, pady=10)
        main_frame.pack_propagate(False)

        # Title
        title = tk.Label(main_frame, text="🔐 User Management", font=("Segoe UI", 20, "bold"),
                         bg="#ffffff", fg="#2c3e50")
        title.pack(pady=(30, 10))

        # Subtitle
        subtitle = tk.Label(main_frame, text="Select your action from the options below", 
                           font=self.subtitle_font, bg="#ffffff", fg="#7f8c8d")
        subtitle.pack(pady=(0, 30))

        # User selection dropdown
        dropdown_frame = tk.Frame(main_frame, bg="#ffffff")
        dropdown_frame.pack(pady=(0, 20))

        tk.Label(dropdown_frame, text="Select User:", font=("Segoe UI", 11, "bold"),
                 bg="#ffffff", fg="#2c3e50").pack(side="left", padx=(0, 10))

        user_var = tk.StringVar()
        users = [f"{u[0]} ({u[3]})" for u in self._load_users_from_excel()]
        user_dropdown = ttk.Combobox(dropdown_frame, textvariable=user_var,
                                     values=users, state="readonly", font=("Segoe UI", 14, "bold"), width=30)
        if users:
            user_dropdown.current(0)
        user_dropdown.pack(side="left")

        # Style the dropdown
        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TCombobox', borderwidth=0, relief="flat",
                        background="#ffffff", fieldbackground="#ffffff")
        style.map('TCombobox', fieldbackground=[('readonly', '#ffffff')])

        # Buttons
        button_frame = tk.Frame(main_frame, bg="#ffffff")
        button_frame.pack(pady=20)

        selected = [None]  # To store the selected user

        def on_submit():
            idx = user_dropdown.current()
            users_list = self._load_users_from_excel()  # Always reload to get the latest
            if idx >= 0 and idx < len(users_list):
                selected[0] = users_list[idx]
                dialog.destroy()
            else:
                messagebox.showerror("Error", "Please select a user.")

        # Re-add watermark at the end
            tk.Label(main_frame, text="Developed by mrgargsir", bg="#ffffff", fg="#6B7280",
                     font=("Segoe UI", 9, "italic"), anchor="se").pack(side="bottom", pady=8, fill="x")

        def on_add_user():
            show_add_user_frame()

        def on_update_password():
            show_update_frame()

        submit_btn = RoundedButton(button_frame, text="🚪 Login", radius=18,
                                  btnbackground="#3498db", btnforeground="#ffffff",
                                  clicked=on_submit, width=140, height=40)
        submit_btn.pack(side="left", padx=10)

        add_btn = RoundedButton(button_frame, text="👥 Add User", radius=18,
                             btnbackground="#e74c3c", btnforeground="#ffffff",
                             clicked=on_add_user, width=140, height=40)
        add_btn.pack(side="left", padx=10)

        update_btn = RoundedButton(button_frame, text="🔑 Update Users", radius=18,
                                  btnbackground="#2ecc71", btnforeground="#ffffff",
                                  clicked=on_update_password, width=180, height=40)
        update_btn.pack(side="left", padx=10)

        # --- Add User Frame ---
        def show_add_user_frame():
            # Remove all widgets, including watermark
            for widget in main_frame.winfo_children():
                widget.pack_forget()
            # --- Add User UI ---
            tk.Label(main_frame, text="👥 Add New User", font=("Segoe UI", 18, "bold"),
                     bg="#ffffff", fg="#2c3e50").pack(pady=(30, 10))

            form_frame = tk.Frame(main_frame, bg="#ffffff")
            form_frame.pack(pady=10, padx=30, anchor="center")

            tk.Label(form_frame, text="Name:", font=self.button_font,
                     bg="#ffffff", fg="#2c3e50", anchor="e", width=12).grid(row=0, column=0, padx=(0, 8), pady=6, sticky="e")
            name_entry = RoundedEntry(form_frame, radius=12, font=("Segoe UI", 14), width=22)
            name_entry.grid(row=0, column=1, pady=8, sticky="w")

            tk.Label(form_frame, text="Username:", font=self.button_font,
                     bg="#ffffff", fg="#2c3e50", anchor="e", width=12).grid(row=1, column=0, padx=(0, 8), pady=6, sticky="e")
            username_entry = RoundedEntry(form_frame, radius=12, font=("Segoe UI", 14), width=22)
            username_entry.grid(row=1, column=1, pady=6, sticky="w")

            tk.Label(form_frame, text="Password:", font=self.button_font,
                     bg="#ffffff", fg="#2c3e50", anchor="e", width=12).grid(row=2, column=0, padx=(0, 8), pady=6, sticky="e")
            password_entry = RoundedEntry(form_frame, radius=12, font=("Segoe UI", 14), width=22)
            password_entry.grid(row=2, column=1, pady=6, sticky="w")

            tk.Label(form_frame, text="User Type:", font=self.button_font,
                     bg="#ffffff", fg="#2c3e50", anchor="e", width=12).grid(row=3, column=0, padx=(0, 8), pady=6, sticky="e")
            user_type_var = tk.StringVar(value="employee")
            radio_frame = tk.Frame(form_frame, bg="#ffffff")
            radio_frame.grid(row=3, column=1, pady=6, sticky="w")
            contractor_radio = tk.Radiobutton(radio_frame, text="Contractor 👷", variable=user_type_var,
                                      value="contractor", font=self.button_font, bg="#ffffff",
                                      fg="#2c3e50", selectcolor="#ffffff", activebackground="#ffffff")
            contractor_radio.pack(side="left", padx=5)
            employee_radio = tk.Radiobutton(radio_frame, text="Employee 👔", variable=user_type_var,
                                            value="employee", font=self.button_font, bg="#ffffff",
                                            fg="#2c3e50", selectcolor="#ffffff", activebackground="#ffffff")
            employee_radio.pack(side="left", padx=5)

            btn_frame = tk.Frame(main_frame, bg="#ffffff")
            btn_frame.pack(pady=16)

            success_label = tk.Label(main_frame, text="", fg="#27ae60", bg="#ffffff", font=("Segoe UI", 12, "bold"))
            success_label.pack(pady=(8, 0))

            def save_user():
                name = name_entry.get().strip()
                username = username_entry.get().strip()
                password = password_entry.get().strip()
                user_type = user_type_var.get()
                # Check for duplicate username
                existing_users = self._load_users_from_excel()
                if any(u[1].lower() == username.lower() for u in existing_users):
                    success_label.config(
                        text="This username already exists, use Update Password field.",
                        fg="#e74c3e"
                    )
                    return
                if all([name, username, password]):
                    self._add_user_to_excel(name, username, password, user_type)
                    name_entry.delete(0, 'end')
                    username_entry.delete(0, 'end')
                    password_entry.delete(0, 'end')
                    success_label.config(text="User added successfully!", fg="#27ae60")
                    main_frame.after(1200, reset_main_frame)
                else:
                    success_label.config(text="Please fill all fields.", fg="#e74c3e")

            save_btn = RoundedButton(btn_frame, text="➕ Add User", radius=16,
                                    btnbackground="#e74c3c", btnforeground="#ffffff",
                                    clicked=save_user, width=120, height=38)
            save_btn.pack(side="left", padx=8)
            back_btn = RoundedButton(btn_frame, text="🔙 Back", radius=16,
                                    btnbackground="#95a5a6", btnforeground="#ffffff",
                                    clicked=lambda: reset_main_frame(), width=100, height=38)
            back_btn.pack(side="left", padx=8)

            # Re-add watermark at the end
            tk.Label(main_frame, text="Developed by mrgargsir", bg="#ffffff", fg="#6B7280",
                     font=("Segoe UI", 9, "italic"), anchor="se").pack(side="bottom", pady=8, fill="x")

        # --- Update Password Frame ---
        def show_update_frame():
            for widget in main_frame.winfo_children():
                widget.pack_forget()
            tk.Label(main_frame, text="🔑 Update Record", font=("Segoe UI", 18, "bold"),
                     bg="#ffffff", fg="#2c3e50").pack(pady=(30, 10))

            form_frame = tk.Frame(main_frame, bg="#ffffff")
            form_frame.pack(pady=30, padx=30, anchor="center")

            # --- User selection dropdown for update ---
            tk.Label(form_frame, text="Select User:", font=self.button_font,
                     bg="#ffffff", fg="#2c3e50", anchor="e", width=12).grid(row=0, column=0, padx=(0, 8), pady=6, sticky="e")
            update_users = self._load_users_from_excel()
            update_usernames = [f"{u[0]} ({u[1]})" for u in update_users]
            update_user_var = tk.StringVar()
            update_user_dropdown = ttk.Combobox(
                form_frame,
                textvariable=update_user_var,
                values=update_usernames,
                state="readonly",
                font=("Segoe UI", 14, "bold"),
                width=22
            )
            if update_usernames:
                update_user_dropdown.current(0)
            update_user_dropdown.grid(row=0, column=1, pady=6, sticky="w")

            # --- New password field ---
            tk.Label(form_frame, text="New Password:", font=self.button_font,
                     bg="#ffffff", fg="#2c3e50", anchor="e", width=12).grid(row=1, column=0, padx=(0, 8), pady=6, sticky="e")
            pass_entry = RoundedEntry(form_frame, radius=12, font=self.button_font, width=18)
            pass_entry.grid(row=1, column=1, pady=6, sticky="w")

            btn_frame = tk.Frame(main_frame, bg="#ffffff")
            btn_frame.pack(pady=16)

            success_label = tk.Label(main_frame, text="", fg="#27ae60", bg="#ffffff", font=("Segoe UI", 12, "bold"))
            success_label.pack(pady=(8, 0))

            def save_pass():
                idx = update_user_dropdown.current()
                if idx < 0 or idx >= len(update_users):
                    success_label.config(text="Please select a user.", fg="#e74c3e")
                    return
                new_password = pass_entry.get().strip()
                if not new_password:
                    success_label.config(text="Please enter a new password.", fg="#e74c3e")
                    return
                # Update password in Excel
                user = update_users[idx]
                wb = openpyxl.load_workbook(self.excel_file_path)
                ws = wb.active
                for row in ws.iter_rows(min_row=2):
                    if (row[0].value == user[0] and row[1].value == user[1] and row[3].value == user[3]):
                        row[2].value = new_password
                        break
                wb.save(self.excel_file_path)
                pass_entry.delete(0, 'end')
                success_label.config(text="Password updated successfully!", fg="#27ae60")
                main_frame.after(1200, reset_main_frame)

            def delete_user():
                idx = update_user_dropdown.current()
                if idx < 0 or idx >= len(update_users):
                    success_label.config(text="Please select a user to delete.", fg="#e74c3e")
                    return
                user = update_users[idx]

                # Remove any previous confirmation widgets
                for widget in btn_frame.winfo_children():
                    if getattr(widget, "_is_confirm", False):
                        widget.destroy()

                # Show confirmation message in success_label
                success_label.config(
                    text=f"Delete user '{user[0]}'?",
                    fg="#e67e22"
                )

                # Yes/No buttons in a frame below the success_label
                confirm_btns = tk.Frame(main_frame, bg="#ffffff")
                confirm_btns.pack(after=success_label, pady=(0, 10))
                confirm_btns._is_confirm = True

                def do_delete():
                    wb = openpyxl.load_workbook(self.excel_file_path)
                    ws = wb.active
                    found = False
                    for row in ws.iter_rows(min_row=2):
                        if (row[0].value == user[0] and row[1].value == user[1] and row[3].value == user[3]):
                            ws.delete_rows(row[0].row, 1)
                            found = True
                            break
                    if found:
                        wb.save(self.excel_file_path)
                        # Refresh dropdown values after deletion
                        new_users = self._load_users_from_excel()
                        new_usernames = [f"{u[0]} ({u[1]})" for u in new_users]
                        update_user_dropdown['values'] = new_usernames
                        if new_usernames:
                            update_user_dropdown.current(0)
                        else:
                            update_user_var.set("")
                        success_label.config(text="User deleted successfully!", fg="#27ae60")
                        main_frame.after(1200, reset_main_frame)
                    else:
                        success_label.config(text="User not found.", fg="#e74c3e")
                    confirm_btns.destroy()

                def cancel_delete():
                    success_label.config(text="", fg="#27ae60")
                    confirm_btns.destroy()

                yes_btn = tk.Button(confirm_btns, text="Yes", font=("Segoe UI", 11, "bold"), bg="#e74c3c", fg="#fff", command=do_delete)
                yes_btn.pack(side="left", padx=8)
                yes_btn._is_confirm = True

                no_btn = tk.Button(confirm_btns, text="No", font=("Segoe UI", 11, "bold"), bg="#95a5a6", fg="#fff", command=cancel_delete)
                no_btn.pack(side="left", padx=8)
                no_btn._is_confirm = True

            save_btn = RoundedButton(btn_frame, text="🔄 Update", radius=16,
                                    btnbackground="#2ecc71", btnforeground="#ffffff",
                                    clicked=save_pass, width=130, height=38)
            save_btn.pack(side="left", padx=8)
            delete_btn = RoundedButton(btn_frame, text="🗑️ Delete User", radius=16,
                                    btnbackground="#e74c3c", btnforeground="#ffffff",
                                    clicked=delete_user, width=170, height=38)
            delete_btn.pack(side="left", padx=8)
            back_btn = RoundedButton(btn_frame, text="🔙 Back", radius=16,
                                    btnbackground="#95a5a6", btnforeground="#ffffff",
                                    clicked=lambda: reset_main_frame(), width=100, height=38)
            back_btn.pack(side="left", padx=8)

            # Re-add watermark at the end
            tk.Label(main_frame, text="Developed by mrgargsir", bg="#ffffff", fg="#6B7280",
                     font=("Segoe UI", 9, "italic"), anchor="se").pack(side="bottom", pady=8, fill="x")

        def reset_main_frame():
            for widget in main_frame.winfo_children():
                widget.pack_forget()
            title.pack(pady=(30, 10))
            subtitle.pack(pady=(0, 30))
            dropdown_frame.pack(pady=(0, 20))
            button_frame.pack(pady=20)
            # Update user list in dropdown
            users = [f"{u[0]} ({u[3]})" for u in self._load_users_from_excel()]
            user_dropdown['values'] = users
            if users:
                user_dropdown.current(len(users) - 1)  # Select the last (newly added) user

        # Watermark
        tk.Label(main_frame, text="Developed by mrgargsir", bg="#ffffff", fg="#6B7280",
                 font=("Segoe UI", 9, "italic"), anchor="se").pack(side="bottom", pady=8, fill="x")

        dialog.deiconify()  # Show the window now that it's ready
        dialog.grab_set()
        dialog.lift()
        dialog.focus_force()
        root.wait_window(dialog)
        root.destroy()

        if selected[0]:
            print(f"[USER] Selected user: {selected[0][0]}")
            return selected[0][1], selected[0][2], selected[0][3]
        else:
            print("[USER] No user selected.")
            return None

    def Captcha_Solver(self):
        """Auto-solve captcha using OCR and return the text."""
        print("[CAPTCHA] Attempting to auto-solve captcha using OCR.")
        driver = self.driver
        wait = self.wait

        try:
            captcha_img = wait.until(EC.presence_of_element_located((By.ID, "imgCaptcha")))
            time.sleep(0.2)  # Ensure image is fully loaded
            print("[CAPTCHA] Captcha image found, extracting source URL.")
            captcha_src = captcha_img.get_attribute("src")
        except Exception as e:
            print(f"[CAPTCHA] Could not find captcha image: {e}")
            return None

        import requests
        from PIL import Image
        import io
        import pytesseract

        # Get the captcha image bytes (use cookies for session)
        session = requests.Session()
        for cookie in driver.get_cookies():
            session.cookies.set(cookie['name'], cookie['value'])

        max_attempts = 3
        captcha_text = ""
        for attempt in range(max_attempts):
            response = session.get(captcha_src)
            img_data = response.content

            # Open image and preprocess for better OCR accuracy
            image = Image.open(io.BytesIO(img_data)).convert("L")  # Grayscale

            # Enhance contrast and sharpness
            image = ImageEnhance.Contrast(image).enhance(2.5)
            image = ImageEnhance.Sharpness(image).enhance(2.0)

            # Apply median filter to reduce noise
            image = image.filter(ImageFilter.MedianFilter(size=3))

            # Threshold (binarize)
            image = image.point(lambda x: 0 if x < 150 else 255, 'L')

            # Convert PIL image to OpenCV format for erosion
            open_cv_image = np.array(image)
            if open_cv_image.ndim == 3:
                open_cv_image = cv2.cvtColor(open_cv_image, cv2.COLOR_RGB2GRAY)

            # Erosion
            kernel = np.ones((2, 2), np.uint8)  # You can adjust kernel size as needed
            eroded = cv2.erode(open_cv_image, kernel, iterations=1)

            # Convert back to PIL Image for pytesseract
            image = Image.fromarray(eroded)

            # OCR: alphanumeric whitelist, expect 6 chars
            try:
                config = '--psm 8 --oem 3 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789'
                captcha_text = pytesseract.image_to_string(image, config=config).strip()
                captcha_text = ''.join(filter(str.isalnum, captcha_text)).upper()
                print(f"[CAPTCHA] OCR attempt {attempt+1} result: {captcha_text}")
                if len(captcha_text) == 6:
                    break  # Success
            except Exception as e:
                print(f"[CAPTCHA] OCR failed: {e}")
                captcha_text = ""
                continue

        # Optionally, show the image and OCR result for manual correction if needed
        problematic_chars = set("S90O8B")  # Add more as needed

        if len(captcha_text) != 6 or any(c in captcha_text for c in problematic_chars):
            print(f"[CAPTCHA] OCR result invalid or contains problematic character(s): {captcha_text}. Asking user for manual input.")
            from PIL import ImageTk
            import tkinter as tk

            root = tk.Tk()
            root.title("Captcha Solver (Manual)")

            # Set window size
            win_w, win_h = 420, 300
            screen_w = root.winfo_screenwidth()
            screen_h = root.winfo_screenheight()
            x = (screen_w // 2) - (win_w // 2)
            y = (screen_h // 2) - (win_h // 2)
            root.geometry(f"{win_w}x{win_h}+{x}+{y}")
            root.attributes("-topmost", True)

            # Show the image
            image = image.resize((240, 80))  # Increase image size for visibility
            photo = ImageTk.PhotoImage(image)
            img_label = tk.Label(root, image=photo)
            img_label.pack(pady=(18, 8))

            # Show OCR result
            # Show OCR result
            ocr_label = tk.Label(root, text=f"OCR Result: {captcha_text}", font=("Segoe UI", 14, "bold"), fg="#2c3e50")
            ocr_label.pack(pady=(0, 12))

            # Frame for character entry boxes
            entry_frame = tk.Frame(root)
            entry_frame.pack(pady=(0, 18))

            # Prepare entry boxes for each character
            char_vars = []
            entries = []
            for i in range(6):
                var = tk.StringVar()
                # Pre-fill with OCR result if available, else empty
                if i < len(captcha_text):
                    var.set(captcha_text[i])
                entry = tk.Entry(entry_frame, textvariable=var, width=2, font=("Segoe UI", 22, "bold"), justify="center")
                entry.grid(row=0, column=i, padx=6)
                char_vars.append(var)
                entries.append(entry)

            # Focus on the first problematic character, or first box if none
            first_problem_idx = None
            for idx, var in enumerate(char_vars):
                if var.get().upper() in problematic_chars:
                    first_problem_idx = idx
                    break
            if first_problem_idx is not None:
                entries[first_problem_idx].focus_set()
                entries[first_problem_idx].icursor("end")
            else:
                entries[0].focus_set()
                entries[0].icursor("end")

            # Tab/Enter navigation
            def on_key(event, idx):
                if event.keysym in ("Tab", "Right"):
                    if idx < 5:
                        entries[idx + 1].focus_set()
                        entries[idx + 1].icursor("end")
                    return "break"
                elif event.keysym == "Left":
                    if idx > 0:
                        entries[idx - 1].focus_set()
                        entries[idx - 1].icursor("end")
                    return "break"
                elif event.keysym == "Return":
                    submit()
                    return "break"

            for idx, entry in enumerate(entries):
                entry.bind("<Tab>", lambda e, i=idx: on_key(e, i))
                entry.bind("<Right>", lambda e, i=idx: on_key(e, i))
                entry.bind("<Left>", lambda e, i=idx: on_key(e, i))
                entry.bind("<Return>", lambda e, i=idx: on_key(e, i))

            # Submit logic
            captcha_value = []

            def submit():
                value = ''.join(var.get().strip().upper()[:1] for var in char_vars)
                if len(value) == 6 and all(c.isalnum() for c in value):
                    captcha_value.append(value)
                    root.destroy()
                else:
                    tk.messagebox.showerror("Error", "Please enter all 6 alphanumeric characters.")

            btn = tk.Button(root, text="Submit", font=("Segoe UI", 12, "bold"), command=submit, bg="#3498db", fg="white")
            btn.pack(pady=(0, 10))

            root.lift()
            root.mainloop()

            if captcha_value:
                captcha_text = captcha_value[0].strip().upper()
                print(f"[CAPTCHA] Manual entry: {captcha_text}")
            else:
                print("[CAPTCHA] No captcha entered.")
                return None

        return captcha_text

    def Login_Process(self, user_tuple=None):
        """Perform login using user details tuple"""
        driver = self.driver
        wait = self.wait

        # Ensure only one tab is open before navigating
        if len(driver.window_handles) > 1:
            for handle in driver.window_handles[1:]:
                driver.switch_to.window(handle)
                driver.close()
            driver.switch_to.window(driver.window_handles[0])

        # Open the login page URL in the current tab
        try:
            driver.get("https://works.haryana.gov.in/HEWP_Login/login.aspx")
            print("[LOGIN] Opened login page URL.")
            self.ensure_window_visible()  # Ensure Chrome window is visible
            time.sleep(0.1)  # Give the page a moment to load
            wait.until(EC.presence_of_element_located((By.ID, "alogin")))  # Wait for login button to appear
        except Exception as e:
            print(f"[LOGIN] Could not open login page: {e}")
            return

        # 1. Click the "Log In" button (by id="alogin")
        try:
            login_btn = driver.find_element(By.ID, "alogin")
            try:
                login_btn.click()
            except Exception:
                driver.execute_script("arguments[0].click();", login_btn)
            print("[LOGIN] Clicked Log In button.")
            time.sleep(0.1)  # Give modal a moment to appear

            wait.until(EC.visibility_of_element_located((By.ID, "imgCaptcha")))  # Wait for captcha to load

        except Exception as e:
            print(f"[LOGIN] Could not click Log In button: {e}")
            return

    # --- SHOW USER SELECTION WINDOW HERE ---
        if user_tuple is None:
            user_tuple = self.user_details()
            if not user_tuple:
                print("[LOGIN] User cancelled selection. Aborting login.")
                return

        username, password, description = user_tuple
        print(f"[LOGIN] Logging in as {username} ({description})")

        # 2. Select the correct radio button based on user type
        try:
            time.sleep(0.2)  # Ensure modal is fully loaded
            if description.lower() == "contractor":
                # Click the label for Contractor (triggers JS and updates UI)
                contractor_label = driver.find_element(By.XPATH, "//label[input[@id='contractorbtn']]")
                contractor_label.click()
                print("[LOGIN] Selected Contractor radio button.")
            elif description.lower() == "employee":
                employee_label = driver.find_element(By.XPATH, "//label[input[@id='employeebtn']]")
                employee_label.click()
                print("[LOGIN] Selected Employee radio button.")
            else:
                print(f"[LOGIN] Unknown user type: {description}")
                return
            time.sleep(0.2)
        except Exception as e:
            print(f"[LOGIN] Could not select user type radio button: {e}")
            return

        # 3. Fill username
        try:
            user_field = driver.find_element(By.ID, "txtuser")
            user_field.clear()
            user_field.send_keys(username)
            print("[LOGIN] Entered username.")
        except Exception as e:
            print(f"[LOGIN] Could not fill username: {e}")
            return

        # 4. Fill password
        try:
            pass_field = driver.find_element(By.ID, "txtpass")
            pass_field.clear()
            pass_field.send_keys(password)
            print("[LOGIN] Entered password.")
        except Exception as e:
            print(f"[LOGIN] Could not fill password: {e}")
            return

        # 5. Solve captcha and fill in <input id="txtCaptcha">
        captcha_field = driver.find_element(By.ID, "txtCaptcha")
        captcha_field.click()  # Focus the field
        captcha_field.clear()
        
        captcha_text = self.Captcha_Solver()
        if not captcha_text:
            print("[LOGIN] No captcha entered.")
            return
        try:
            
            captcha_field.send_keys(captcha_text)
            print("[LOGIN] Entered captcha.")
        except Exception as e:
            print(f"[LOGIN] Could not fill captcha: {e}")
            return

        # 6. Click the final submit/login button (id="btnLogin")
        try:
            submit_btn = driver.find_element(By.ID, "btnLogin")
            try:
                submit_btn.click()
            except Exception:
                driver.execute_script("arguments[0].click();", submit_btn)
            print("[LOGIN] Clicked Login/Submit button.")
            

        except Exception as e:
            print(f"[LOGIN] Could not click submit button: {e}")
            return

        # 7. Check for successful login
        try:
            # Wait briefly for any alert to appear
            time.sleep(1)
            alert = None
            try:
                alert = driver.switch_to.alert
                alert_text = alert.text
                status = alert_text
                alert.accept()  # Optionally close the alert
                print(f"[LOGIN] Chrome notification found: {alert_text}")
            except Exception:
                # No alert present
                status = "Log in Success"
                print("[LOGIN] No Chrome notification found, assuming login success.")
        except Exception as e:
            status = f"Login status check failed: {e}"
            print(f"[LOGIN] Error checking login status: {e}")

        self.show_notification(status)
        print(status)

        # Move the selected user one row up in Excel if login was successful
        if status == "Log in Success" and username and password and description:
            try:
                wb = openpyxl.load_workbook(self.excel_file_path)
                ws = wb.active
                users = list(ws.iter_rows(min_row=2, values_only=True))
                # Find the index of the selected user
                idx = next((i for i, row in enumerate(users)
                            if row[1] == username and row[2] == password and row[3] == description), None)
                if idx is not None and idx > 0:
                    # Swap with the row above
                    users[idx], users[idx-1] = users[idx-1], users[idx]
                    # Rewrite all user rows
                    for i, row in enumerate(users, start=2):
                        for j, value in enumerate(row, start=1):
                            ws.cell(row=i, column=j, value=value)
                    wb.save(self.excel_file_path)
                    print(f"[LOGIN] Moved user '{username}' up in preference list.")
            except Exception as e:
                print(f"[LOGIN] Could not update user preference order: {e}")

    def process_item(self):
        """Main workflow with window management"""
        print("[PROCESS] Starting main process workflow...")
        try:
            self.Login_Process()  # No need to call self.user_details() here
            # self.minimize_chrome_window() 
            return True
        except Exception as e:
            print(f"[PROCESS] Processing failed: {str(e)}")
            messagebox.showerror("Error", f"Processing failed:\n{str(e)}")
            return False
        finally:
            print("[PROCESS] Cleaning up resources...")
            # Only close resources if user actually logged in
            # self.close()

    def close(self):
        """Clean up resources"""
        print("[CLOSE] Closing resources...")
        if hasattr(self, 'driver'):
            try:
                #self.driver.quit()
                print("[CLOSE] skkip WebDriver closed.")
            except Exception as e:
                print(f"[CLOSE] Error closing WebDriver: {e}")

        if hasattr(self, 'notification_popup') and self.notification_popup is not None:
            try:
                self.notification_popup.destroy()
                print("[CLOSE] Notification popup destroyed.")
            except Exception as e:
                print(f"[CLOSE] Error destroying notification popup: {e}")

        if self.notification_root:
            try:
                self.notification_root.destroy()
                print("[CLOSE] Notification window destroyed.")
            except Exception as e:
                print(f"[CLOSE] Error destroying notification window: {e}")

    def minimize_chrome_window(self):
        try:
            chrome_windows = gw.getWindowsWithTitle("Chrome")
            if chrome_windows:
                chrome_windows[0].minimize()
        except Exception as e:
            print(f"Could not minimize Chrome: {e}")

def run_automation():
    """Run the automation process"""
    print("[RUN] Starting automation...")
    mrgargsir = HEWPLogin()
    try:
        mrgargsir.process_item()
    finally:
        mrgargsir.close()
    print("[RUN] Automation finished.")

if __name__ == "__main__":
    run_automation()
    sys.exit(0)  # Ensures the script exits cleanly